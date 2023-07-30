#!/usr/bin/env python3
# coding: utf-8
import csv
import asyncio
import glob
import json
import os
import time
import graphviz
from .libxx import build_bot_tpl, bsc_relation_read
from .mistapi import MistAcquireDat
import pandas as pd
from pandas import DataFrame
import openpyxl as ox

ALLSIDENODE = False


class MistAnalysis:
    """
    https://graphviz.org/doc/info/shapes.html

    """

    def __init__(self):
        self.folder = "data/mist"
        self.handle_address = ""
        self.metadata = {
            "LINK": {},
            "IDS": {},
            "USE_NODE": [],
            "FROM_LIST": []
        }
        self.rendered = False
        self.scope = 500
        self.dot = build_bot_tpl(
            "collection",
            self.scope
        )
        self.api = MistAcquireDat()
        self.edges = 0

    def setThreadHoldUSD(self, n):
        self.scope = n
        return self

    def start(self):
        file_pattern = "*.json"  # Replace with your specific file name pattern
        # Use glob to search for files with the specified name pattern
        # search = os.path.join(os.path.dirname(__file__), self.folder, file_pattern)
        search = os.path.join(self.folder, file_pattern)
        file_list = glob.glob(search)
        # Loop over each file and perform some operation
        for file_path in file_list:
            file_name = os.path.basename(file_path)
            self._inside(file_path, self.scope)

        self.end()

    def _inside(self, file, scope):
        allsidenote = ALLSIDENODE
        with open(file, newline='') as f:
            self.rf = json.loads(f.read())
            # render_node = True
            if "graph_dic" not in self.rf:
                return
            text = f"下图显示了所有被 {scope} USDT 或以下忽略的记录交易"

            with self.dot.subgraph(name='legend') as c:
                c.attr(color='blue')
                c.node_attr['style'] = 'filled'
                c.attr(label=text)

            nodes = self.rf["graph_dic"]["node_list"]
            edges = self.rf["graph_dic"]["edge_list"]

            for v in nodes:
                _font_colr = "blue"
                _id = v["id"]
                _add = v["addr"]

                if "shape" in v:
                    _shape = v["shape"]
                    if _shape == "circularImage":
                        _shape = "box"
                    print(_shape)
                else:
                    _shape = "box"

                if _add not in self.metadata:
                    self.metadata[_add] = _id
                else:
                    old_id = self.metadata[_add]
                    self.metadata["LINK"][_id] = old_id
                    _id = old_id
                    # render_node = False

                if "..." in v["label"]:
                    _lab = v["addr"]
                    _color = "deepskyblue1"

                    if "star" in _shape:
                        side_note = self.api.overview(_add)
                        _lab = _lab + side_note
                        _font_colr = "white"
                        _shape = "folder"

                else:
                    _color = "firebrick1"
                    _lab = v["label"] + ":\n" + v["addr"]

                    if "star" in _shape:
                        side_note = self.api.overview(_add)
                        _lab = _lab + side_note
                        _font_colr = "white"
                        _shape = "folder"

                    if "fire" in _color and allsidenote:
                        side_note = self.api.overview(_add)
                        _lab = _lab + side_note

                if "color" in v:
                    _color = v["color"]

                self.metadata["IDS"][_id] = {
                    "shape": _shape,
                    "fillcolor": _color,
                    "style": "filled",
                    "fontcolor": _font_colr,
                    "label": _lab
                }

            for v in edges:
                if v["val"] < scope:
                    continue
                if "color" in v:
                    _color = v["color"]["color"]
                else:
                    _color = "red"
                count = len(v["tx_hash_list"])
                __label = v["label"]

                self.dot.edge(
                    self.getId(v["from"]),
                    self.getId(v["to"]),
                    color=_color,
                    label=f"{count}笔,{__label}"
                )


                self.edges += 1

            for _id in self.metadata["USE_NODE"]:
                # if render_node:
                ob = self.metadata["IDS"][_id]
                self.dot.node(
                    _id,
                    shape=ob["shape"],
                    fillcolor=ob["fillcolor"],
                    style="filled",
                    fontcolor=ob["fontcolor"],
                    label=ob["label"]
                )

    def getId(self, _idx):
        if _idx in self.metadata["LINK"]:
            use_alternative = self.metadata["LINK"][_idx]
        else:
            use_alternative = _idx

        if use_alternative not in self.metadata["USE_NODE"]:
            self.metadata["USE_NODE"].append(use_alternative)

        return use_alternative

    def end(self):
        self.dot.render(directory='data/charts').replace('\\', '/')

    def _readDF(self, excel_file: str) -> DataFrame:
        data = pd.read_excel(excel_file)
        df = pd.DataFrame(data, columns=self._excel_header)
        return df

    def _update(self, excel_file: str) -> "MistAnalysis":

        self._excel_header = [
            # 2         3         4       5        6         7     8          9            10
            "address", "start", "end", "income", "outflow", "net", "tx", "spent count", "contribution"
        ]

        try:
            _df = self._readDF(excel_file)
            wb = ox.load_workbook(excel_file)
            ws = wb["Sheet1"]

            startcol: int = 0
            startrow: int = 0
            key = ""
            original_rows = _df.shape[0]
            original_cols = _df.shape[1]
            original_cols = len(self._excel_header)
            original_rows = len(self.metadata["FROM_LIST"])
            for row in range(0, original_rows):
                if row >= startrow:
                    rT = row + 2
                    package_info = self.metadata["FROM_LIST"][row]["info"]
                    address = self.metadata["FROM_LIST"][row]["address"]
                    spent_count = self.metadata["FROM_LIST"][row]["spent_count"]
                    contribute = self.metadata["FROM_LIST"][row]["contribute"]

                    ws.cell(row=rT, column=2).value = address
                    ws.cell(row=rT, column=3).value = package_info["first_tx_time"]
                    ws.cell(row=rT, column=4).value = package_info["last_tx_time"]
                    ws.cell(row=rT, column=5).value = package_info["total_received"]
                    ws.cell(row=rT, column=6).value = package_info["total_spent"]
                    ws.cell(row=rT, column=7).value = package_info["balance"]
                    ws.cell(row=rT, column=8).value = package_info["tx_count"]
                    ws.cell(row=rT, column=9).value = spent_count
                    ws.cell(row=rT, column=10).value = contribute

            """
            for row in range(0, original_rows):  # For each row in the dataframe
                for col in range(0, original_cols):  # For each column in the dataframe
                    if row >= startrow:
                        val_old = _df.iat[row, col]
                        key = self._excel_header[col]
                        if col == 0:
                            rT = row + 2
                            ws.cell(row=rT, column=2).value = pass_discord
                            ws.cell(row=rT, column=4).value = pass_email
                            ws.cell(row=rT, column=5).value = token_p
                            ws.cell(row=rT, column=9).value = banned
                            ws.cell(row=rT, column=10).value = last_operation
                            ws.cell(row=rT, column=11).value = bind_phone
                            # print(f"Pop =|> {rT} {val_old} {k1v1} {k1v2} {k1v3}")
            """

            wb.save(excel_file)

        except FileNotFoundError:
            pass

        return self


def graph_building_bsc_relations(project_name: str):
    dot = graphviz.Digraph(
        project_name,
        comment='Sun',
        engine='dot',
        format='pdf'
    )
    dot.attr(
        ordering='out',
        k='2.2',
        overlap='prism0',
        rankdir='LR',
        size='1000,250'
    )
    # dot.attr('node', shape='doublecircle')
    # dot.format = 'pdf'
    # dot.engine = 'dot'
    # dot.engine = 'neato'
    dot = bsc_relation_read(dot)
    dot.render(directory='data/charts').replace('\\', '/')
